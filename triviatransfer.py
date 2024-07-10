import openpyxl

def create_sheet_if_not_exists(workbook, sheet_title):
  """
  This function checks if a sheet with the given title exists in the workbook.
  If not, it creates a new sheet with that title.

  Args:
      workbook: An openpyxl.Workbook object.
      sheet_title (str): The title of the sheet to check and potentially create.
  """
  if sheet_title not in workbook.sheetnames:
    # Sheet doesn't exist, create a new one
    new_sheet = workbook.create_sheet(sheet_title)
    print(f"Sheet '{sheet_title}' created.")
    return
  else:
    return

def write_to_next_row(workbook, sheet_name, data_to_add):
  """
  This function writes data to the next available row in a specified sheet.

  Args:
      workbook: An openpyxl.Workbook object.
      sheet_name (str): The name of the sheet to write data to.
      data_to_add (list): A list containing the data to write to each column in the row.
  """
  sheet = workbook[sheet_name]

  # Find the first empty row (starting from row 2 as row 1 is likely the header)
  row_to_write = 2
  while sheet.cell(row=row_to_write, column=1).value is not None:
    row_to_write += 1

  # Write the data to the next available row
  for col, value in enumerate(data_to_add, start=1):
    sheet.cell(row=row_to_write, column=col).value = value
  
  print(f"Question #{data_to_add[0]} added to {data_to_add[1]}")


old_workbook = openpyxl.load_workbook("Trivia-Printable.xlsx")
old_sheet = old_workbook["Trivia"]

wb = openpyxl.Workbook()
wb.save("trivia.xlsx")
new_workbook = openpyxl.load_workbook("trivia.xlsx")



identifier = 1

for row in old_sheet: 
  if row[0].value != None:
    category = row[0].value
    question = row[1].value
    answer = row[2].value

    question_data = [identifier, category, question, answer]

    create_sheet_if_not_exists(new_workbook, category)
    write_to_next_row(new_workbook, category, question_data)
    identifier += 1
  else:
    print("Empty Cell")


  if row[4].value != None:
    category = row[4].value
    question = row[5].value
    answer = row[6].value

    question_data = [identifier, category, question, answer]

    create_sheet_if_not_exists(new_workbook, category)
    write_to_next_row(new_workbook, category, question_data)
    identifier += 1
  else:
    print("Empty Cell")
  
  if row[8].value != None:
    category = row[8].value
    question = row[9].value
    answer = row[10].value

    question_data = [identifier, category, question, answer]

    create_sheet_if_not_exists(new_workbook, category)
    write_to_next_row(new_workbook, category, question_data)
    identifier += 1
  else:
    print("Empty Cell")

  if identifier % 1000 == 0:
    new_workbook.save("trivia.xlsx")
    print("Workbook saved.")

new_workbook.save("trivia.xlsx")

